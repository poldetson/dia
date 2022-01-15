import sys
#import os
import csv
import openpyxl
import jaconv

#print(sys.version)
#print(sys.path)
#os.getcwd()

# open stations csv file & read it
stations = []
csvfile = open("station.txt", encoding="utf-8")
for row in csv.reader(csvfile, skipinitialspace=True):
    station = [row[0], int(row[1])]
    stations.append(station)
#print("station csv:")
#print(stations)

# open excel file
wb = openpyxl.load_workbook('aidia2.xlsx')
sheets = wb.get_sheet_names()
wsheet = wb.get_sheet_by_name('Table 1')

# open csv file to write
wcsvfile = open('converted.csv', 'w', newline="")
writer = csv.writer(wcsvfile)
#writer.writerow([0, 1, 2])

# goto next train data point
row = 2
column = 3

## write train name 
# get train name & convert zenkaku -> hankaku, delete space
trains = []
while wsheet.cell(row, column).value != None:
    train_name = jaconv.z2h(wsheet.cell(row, column).value, digit=True, ascii=True).replace(" ", "")
    #print(train_name)
    trains.append(train_name)
    column = column + 1

# write station no.
row = 7
column = 1
station_name = ""
points = []
for i in range(24):
    last_station_name = station_name
    station_name = wsheet.cell(row + i, column).value
    if station_name == None:
        station_name = last_station_name
    else:
        station_name = jaconv.z2h(station_name, digit=True, ascii=True).replace(" ", "")
    #print(station_name)
    station_no = 0
    for station in stations:
        if (station_name == station[0]):
            #print("station no.:" + str(station_no))
            points.append(station_no)
            break
        else:
            station_no = station_no + 1

# write train name
writer.writerow([trains[0]])

# write time at station
row = 7
column = 3
i = 0
j = 0
for train_name in trains:
   for point in points:
      tmp = wsheet.cell(row + i, column + j).value
      if tmp != None:
         time_at_station = jaconv.z2h(tmp, digit=True, ascii=True).replace(" ", "")
         formatted_time = "{:0>5}".format(time_at_station)
         #print("{:0>5}".format(time_at_station))
         writer.writerow([point, formatted_time])
#      else:
         #print("None")
      i = i + 1
   # write end statement
   writer.writerow(["END"])
   j = j + 1

# close csv file